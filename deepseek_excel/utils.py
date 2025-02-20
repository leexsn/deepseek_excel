from openpyxl import load_workbook

def read_excel(filename, cell_range=None):
    """
    读取Excel文件中的数据。
    
    :param filename: Excel文件路径
    :param cell_range: 要读取的单元格范围，默认为整个工作表
    :return: 包含单元格值的二维列表
    """
    try:
        # 加载工作簿时启用只读模式以提高性能
        wb = load_workbook(filename=filename, read_only=True)
        sheet = wb.active
        
        if cell_range is None:
            # 如果没有指定单元格范围，则读取整个工作表
            data = [[cell.value for cell in row] for row in sheet.iter_rows()]
        else:
            # 解析单元格区域并提取值，保持原始行列结构
            data = [[cell.value for cell in row] for row in sheet[cell_range]]
        
        return data
    except Exception as e:
        print(f"读取Excel文件错误: {e}")
        return []
    finally:
        # 确保工作簿被正确关闭
        if 'wb' in locals():
            wb.close()

# 示例用法
if __name__ == "__main__":
    file_path = "example.xlsx"
    cell_range = "A1:B2"  # 例如，读取从A1到B2的单元格范围
    data = read_excel(file_path, cell_range)
    for row in data:
        print(row)